#!/usr/bin/env python3
"""Export a video-master production workbook."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from delivery_paths import (
    HANDOFF,
    IMAGE_PROMPTS,
    OVERVIEW_PNG,
    PREVIEW_MP4,
    VIDEO_PROMPTS,
    WORKBOOK,
)


SHOT_COLUMNS = [
    "shot_id",
    "start",
    "end",
    "duration_seconds",
    "beat",
    "purpose",
    "visual_action",
    "framing",
    "camera_angle",
    "lens_depth",
    "camera_movement",
    "lighting",
    "audio_refs",
    "onscreen_text_refs",
    "continuity_notes",
]

VOICEOVER_COLUMNS = ["id", "start", "end", "text", "tone", "speed", "pause_after_ms"]

FILE_ROWS = [
    ("Creative brief", "brief/creative_brief.md"),
    ("Spec lock", "brief/spec_lock.md"),
    ("Creative strategy", "strategy/creative_strategy.md"),
    ("Rhythm map", "strategy/rhythm_map.md"),
    ("Script", "script/script.md"),
    ("Shot list", "storyboard/shot_list.md"),
    ("Video prompts", "prompts/video_prompts.md"),
    ("Copy-ready video prompts", VIDEO_PROMPTS.as_posix()),
    ("Copy-ready image prompts", IMAGE_PROMPTS.as_posix()),
    ("Storyboard overview PNG", OVERVIEW_PNG.as_posix()),
    ("Animatic preview", PREVIEW_MP4.as_posix()),
    ("Handoff", HANDOFF.as_posix()),
]


def read_json(path: Path, default: Any) -> Any:
    if not path.is_file():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def setup_sheet(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="E9EEF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"


def autosize(sheet) -> None:
    for column in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 48)


def write_shots(workbook: Workbook, project: Path) -> None:
    sheet = workbook.active
    sheet.title = "Shots"
    setup_sheet(sheet, SHOT_COLUMNS)
    for shot in read_json(project / "storyboard" / "shot_list.json", []):
        sheet.append([stringify(shot.get(column)) for column in SHOT_COLUMNS])
    autosize(sheet)


def write_voiceover(workbook: Workbook, project: Path) -> None:
    sheet = workbook.create_sheet("Voiceover")
    setup_sheet(sheet, VOICEOVER_COLUMNS)
    for row in read_json(project / "audio" / "tts_lines.json", []):
        sheet.append([stringify(row.get(column)) for column in VOICEOVER_COLUMNS])
    autosize(sheet)


def write_files(workbook: Workbook, project: Path) -> None:
    sheet = workbook.create_sheet("Files")
    setup_sheet(sheet, ["label", "path", "exists"])
    for label, relative in FILE_ROWS:
        sheet.append([label, relative, "yes" if (project / relative).exists() else "no"])
    autosize(sheet)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Export a production workbook for a video-master project.")
    parser.add_argument("project", type=Path, help="Path to a video-master project directory")
    parser.add_argument("-o", "--output", type=Path, help="Workbook output path")
    args = parser.parse_args(argv)

    project = args.project.resolve()
    if not project.is_dir():
        print(f"ERROR: project directory does not exist: {project}")
        return 2

    workbook = Workbook()
    write_shots(workbook, project)
    write_voiceover(workbook, project)
    write_files(workbook, project)
    output = args.output or project / WORKBOOK
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(f"Workbook: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
