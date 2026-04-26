import json
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXPORTER = ROOT / "skills" / "video-master" / "scripts" / "export_production_workbook.py"


def write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


class ExportProductionWorkbookTest(unittest.TestCase):
    def test_exports_workbook_with_shots_audio_and_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp) / "project"
            write(
                project / "storyboard" / "shot_list.json",
                json.dumps(
                    [
                        {
                            "shot_id": "S01",
                            "start": "00:00",
                            "end": "00:03",
                            "duration_seconds": 3,
                            "beat": "Hook",
                            "purpose": "抓住注意力",
                            "visual_action": "产品出现",
                        }
                    ],
                    ensure_ascii=False,
                ),
            )
            write(
                project / "audio" / "tts_lines.json",
                json.dumps([{"id": "VO01", "start": "00:00", "end": "00:03", "text": "一滴唤醒水光"}], ensure_ascii=False),
            )
            write(project / "prompts" / "video_prompts.md", "# Video Prompts\n\n## S01\n画面：产品出现\n")
            write(project / "最终交付" / "00_使用说明.md", "# Handoff\n")

            result = subprocess.run(
                ["python3", str(EXPORTER), str(project)],
                cwd=ROOT,
                text=True,
                capture_output=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            workbook_path = project / "最终交付" / "06_制作总表" / "制作总表.xlsx"
            self.assertTrue(workbook_path.is_file())
            workbook = load_workbook(workbook_path)
            self.assertEqual({"Shots", "Voiceover", "Files"}.issubset(set(workbook.sheetnames)), True)
            self.assertEqual(workbook["Shots"]["A2"].value, "S01")
            self.assertEqual(workbook["Voiceover"]["D2"].value, "一滴唤醒水光")


if __name__ == "__main__":
    unittest.main()
